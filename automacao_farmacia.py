"""
Automação para coleta de descrições e imagens de produtos farmacêuticos.
Sites: Drogaria São Paulo (API VTEX), Panvel (SPA Angular) e Beleza na Web (SPA - usa Playwright)
"""

import os
import re
import html
import json
import time
import hashlib
import requests
import openpyxl
from io import BytesIO
from pathlib import Path
from PIL import Image
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout


# ─── CONFIGURAÇÕES ────────────────────────────────────────────────────────────
PLANILHA_ENTRADA   = "produtos.xlsx"  # <-- nome da sua planilha
COLUNA_EAN         = "A"             # coluna com os EANs
COLUNA_DESC_DSP    = "B"             # coluna Descrição Drogaria SP
COLUNA_DESC_PANVEL = "C"             # coluna Descrição Panvel
COLUNA_DESC_BLZ    = "D"             # coluna Descrição Beleza na Web   <-- NOVO
PASTA_IMAGENS      = "imagens"       # pasta raiz para salvar imagens
MAX_IMAGENS        = 10              # limite de imagens únicas por EAN

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9",
}


# ─── UTILITÁRIOS ──────────────────────────────────────────────────────────────

def hash_imagem(conteudo: bytes) -> str:
    return hashlib.md5(conteudo).hexdigest()


def baixar_imagem(url: str, session: requests.Session) -> bytes | None:
    try:
        r = session.get(url, timeout=15, headers=HEADERS)
        if r.status_code == 200 and "image" in r.headers.get("content-type", ""):
            return r.content
    except Exception as e:
        print(f"    ✗ Erro ao baixar {url}: {e}")
    return None


def salvar_imagens_unicas(ean: str, urls: list, session: requests.Session) -> int:
    """
    Baixa e salva até MAX_IMAGENS imagens únicas (por hash MD5).
    Pasta: {PASTA_IMAGENS}/{EAN}/
    Nomes: EAN - 1.jpg, EAN - 2.jpg, ...
    """
    pasta_ean = Path(PASTA_IMAGENS) / str(ean)
    pasta_ean.mkdir(parents=True, exist_ok=True)

    hashes_vistos: set = set()
    contador = 1

    for url in urls:
        if contador > MAX_IMAGENS:
            break
        conteudo = baixar_imagem(url, session)
        if conteudo is None:
            continue
        h = hash_imagem(conteudo)
        if h in hashes_vistos:
            print(f"    ⚠ Duplicada ignorada: {url}")
            continue
        hashes_vistos.add(h)

        ext = url.split("?")[0].rsplit(".", 1)[-1].lower()
        ext = ext if ext in {"jpg", "jpeg", "png", "webp", "gif"} else "jpg"

        nome = pasta_ean / f"{ean}-{contador}.{ext}"
        try:
            img = Image.open(BytesIO(conteudo))
            if img.mode in ("P", "RGBA", "LA"):
                img = img.convert("RGBA").convert("RGB")
            else:
                img = img.convert("RGB")
            img = img.resize((1000, 1000), Image.LANCZOS)
            save_ext = "jpeg" if ext in {"jpg", "jpeg"} else ext
            img.save(nome, format=save_ext.upper() if save_ext != "jpeg" else "JPEG")
        except Exception as e:
            print(f"    ⚠ Erro ao redimensionar, salvando original: {e}")
            nome.write_bytes(conteudo)
        print(f"    ✓ Salva (1000x1000): {nome.name}")
        contador += 1

    return contador - 1


def limpar_html(texto: str) -> str:
    if not texto:
        return texto
    texto = re.sub(r'<[^>]+>', '', texto)
    texto = html.unescape(texto)
    return re.sub(r'\s+', ' ', texto).strip()


def col_letter_to_index(letter: str) -> int:
    letter = letter.upper().strip()
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


# ─── DROGARIA SÃO PAULO (API VTEX) ────────────────────────────────────────────

DSP_BASE = "https://www.drogariasaopaulo.com.br"


def coletar_dsp(ean: str, session: requests.Session) -> tuple:
    """Usa a API VTEX nativa — rápido, sem JavaScript."""
    print(f"  [DSP] Buscando EAN {ean}...")
    url = (
        f"{DSP_BASE}/api/catalog_system/pub/products/search"
        f"?fq=alternateIds_Ean:{ean}&_from=0&_to=0"
    )
    try:
        r = session.get(url, headers=HEADERS, timeout=20)
        data = r.json()
    except Exception as e:
        print(f"  ✗ DSP: Erro na API: {e}")
        return "", []

    if not data:
        print(f"  – DSP: Não encontrado para EAN {ean}")
        return "", []

    produto = data[0]
    todas_imgs = []
    for sku in produto.get("items", []):
        for img in sku.get("images", []):
            img_url = img.get("imageUrl", "")
            if img_url and img_url not in todas_imgs:
                todas_imgs.append(img_url)

    descricao = produto.get("description") or produto.get("productName", "")
    print(f"  [DSP] ✓ {produto.get('productName', '')} | {len(todas_imgs)} imagens")
    return descricao, todas_imgs


# ─── PANVEL (Playwright — SPA Angular) ────────────────────────────────────────

PANVEL_BUSCA = "https://www.panvel.com/panvel/buscarProduto.do?termoPesquisa={ean}"


def coletar_panvel(ean: str, page) -> tuple:
    """
    Usa Playwright para renderizar o JavaScript da Panvel.
    """
    print(f"  [Panvel] Buscando EAN {ean}...")
    try:
        # ── 1. Página de busca ─────────────────────────────────────────────
        page.goto(
            f"https://www.panvel.com/panvel/buscarProduto.do?termoPesquisa={ean}",
            wait_until="domcontentloaded",
            timeout=30000
        )

        # Aguarda o Angular renderizar os links de produto
        try:
            page.wait_for_selector('a[href*="/p-"]', state="visible", timeout=15000)
        except PwTimeout:
            print(f"  – Panvel: Produto não encontrado para EAN {ean}")
            return "", []

        # ── 2. Pega o link do primeiro produto ─────────────────────────────
        produto_url = page.locator('a[href*="/p-"]').first.get_attribute("href")
        if not produto_url:
            print(f"  – Panvel: Link não encontrado para EAN {ean}")
            return "", []

        if produto_url.startswith("/"):
            produto_url = "https://www.panvel.com" + produto_url

        # ── 3. Abre a página do produto ────────────────────────────────────
        time.sleep(0.5)
        page.goto(produto_url, wait_until="domcontentloaded", timeout=30000)

        # Espera o h1 do produto ficar visível (confirma que a página carregou)
        try:
            page.wait_for_selector("h1.product-title", state="visible", timeout=15000)
        except PwTimeout:
            try:
                page.wait_for_selector("h1", state="visible", timeout=10000)
            except PwTimeout:
                print(f"  ⚠ Panvel: Página do produto demorou para carregar ({produto_url})")

        # ── 4. Lê o JSON-LD diretamente do DOM ─────────────────────────────
        ld_json_list = page.evaluate("""() => {
            const scripts = document.querySelectorAll('script[type="application/ld+json"]');
            return Array.from(scripts).map(el => {
                try { return JSON.parse(el.textContent); } catch(e) { return null; }
            }).filter(Boolean);
        }""")

        produto_ld = next((d for d in ld_json_list if d.get("@type") == "Product"), None)

        if produto_ld:
            images = produto_ld.get("image", [])
            if isinstance(images, str):
                images = [images]
            desc = produto_ld.get("description", "")
            nome = produto_ld.get("name", "")
            print(f"  [Panvel] ✓ {nome} | {len(images)} imagens")
            return desc, images

        # ── 5. Fallback: extrair do DOM ────────────────────────────────────
        titulo = page.locator("h1").first.text_content(timeout=5000).strip()
        imgs = page.evaluate("""() =>
            [...new Set(
                Array.from(document.querySelectorAll("img[src*='staticpanvel']"))
                    .map(e => e.src)
            )]
        """)
        desc_items = page.locator(".destaques-produto li").all_text_contents()
        desc = " | ".join(i.strip() for i in desc_items if i.strip())
        print(f"  [Panvel] ✓ (fallback) {titulo} | {len(imgs)} imagens")
        return desc, imgs

    except Exception as e:
        print(f"  ✗ Panvel: Erro para EAN {ean}: {e}")
        return "", []


# ─── BELEZA NA WEB (Playwright + bypass Akamai) ───────────────────────────────

BLZ_BASE = "https://www.belezanaweb.com.br"

def _blz_normalizar_url(url: str) -> str:
    """
    As imagens do Beleza na Web vêm do Cloudinary, mas o DOM entrega o caminho
    de forma RELATIVA (ex: 'w_800/v1/imagens/product/...') ou absoluta.
    Esta função:
      1. Garante o domínio absoluto do Cloudinary.
      2. Substitui a transformação de tamanho por uma de alta qualidade.
    """
    PREFIX = "https://res.cloudinary.com/beleza-na-web/image/upload/"
    TRANSFORM = "w_1500,f_auto,fl_progressive,q_auto:best"

    if not url:
        return url

    # Caso já seja absoluta (cloudinary), apenas troca a transformação
    if url.startswith("http") and "res.cloudinary.com" in url and "/v1/" in url:
        return re.sub(r"/upload/.*?/v1/", f"/upload/{TRANSFORM}/v1/", url)

    # Caso seja relativa: pega só a partir de '/v1/' e remonta com prefixo + transform
    if "/v1/" in url:
        resto = "v1/" + url.split("/v1/", 1)[1]
        return f"{PREFIX}{TRANSFORM}/{resto}"

    # Fallback: se vier sem esquema mas com 'imagens/product', prefixa o domínio
    if url.startswith("imagens/") or "/imagens/product/" in url:
        caminho = url[url.index("imagens/"):] if "imagens/" in url else url
        return f"{PREFIX}{TRANSFORM}/v1/{caminho}"

    return url

_blz_sessao_aquecida = {"ok": False}   # controla o "warm-up" da home (1x por execução)


def _blz_aquecer_sessao(page):
    """
    O Akamai exige cookies de sensor (bm_sv, _abck, etc.) gerados ao visitar
    a home com um navegador 'real'. Sem isso, /busca retorna 'Access Denied'.
    Visita a home uma vez para estabelecer a sessão.
    """
    if _blz_sessao_aquecida["ok"]:
        return
    try:
        page.goto(BLZ_BASE + "/", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2500)   # dá tempo do Akamai injetar os cookies
        _blz_sessao_aquecida["ok"] = True
    except Exception as e:
        print(f"  ⚠ BLZ: falha ao aquecer sessão: {e}")


def coletar_belezanaweb(ean: str, page) -> tuple:
    print(f"  [BLZ] Buscando EAN {ean}...")
    try:
        # ── 0. Aquece a sessão (home) antes de buscar ──────────────────────
        _blz_aquecer_sessao(page)

        # ── 1. Busca (redireciona para o produto se houver match) ──────────
        page.goto(
            f"{BLZ_BASE}/busca?q={ean}",
            wait_until="domcontentloaded",
            timeout=30000
        )

        # Detecta bloqueio do Akamai e tenta reaquecer 1x
        if "Access Denied" in (page.title() or "") or "Access Denied" in page.content()[:500]:
            print(f"  ⚠ BLZ: Akamai bloqueou — reaquecendo sessão e tentando de novo...")
            _blz_sessao_aquecida["ok"] = False
            _blz_aquecer_sessao(page)
            page.wait_for_timeout(3000)
            page.goto(f"{BLZ_BASE}/busca?q={ean}",
                      wait_until="domcontentloaded", timeout=30000)
            if "Access Denied" in (page.title() or ""):
                print(f"  ✗ BLZ: Bloqueado pelo Akamai para EAN {ean}")
                return "", []

        # Se continuou na listagem, pega o primeiro produto
        if "/busca" in page.url:
            try:
                page.wait_for_selector('a[href]', state="visible", timeout=8000)
                link = page.evaluate("""() => {
                    const a = Array.from(document.querySelectorAll('a[href]'))
                        .map(e => e.getAttribute('href'))
                        .find(h => h && /^\\/[a-z0-9-]+-\\d{3,}/i.test(h) && !h.includes('busca'));
                    return a || null;
                }""")
            except PwTimeout:
                link = None
            if not link:
                print(f"  – BLZ: Produto não encontrado para EAN {ean}")
                return "", []
            if link.startswith("/"):
                link = BLZ_BASE + link
            page.goto(link, wait_until="domcontentloaded", timeout=30000)

        # ── 2. Espera o produto carregar ───────────────────────────────────
        try:
            page.wait_for_selector("h1", state="visible", timeout=15000)
        except PwTimeout:
            print(f"  ⚠ BLZ: Página demorou para carregar ({page.url})")

        # ── 3. Extrai JSON-LD (ProductGroup) + imagens pelo SKU ────────────
        dados = page.evaluate("""() => {
            const scripts = document.querySelectorAll('script[type="application/ld+json"]');
            let pg = null;
            for (const el of scripts) {
                try {
                    let d = JSON.parse(el.textContent);
                    const arr = Array.isArray(d) ? d : [d];
                    for (const it of arr) {
                        if (it && (it['@type'] === 'ProductGroup' || it['@type'] === 'Product')) pg = it;
                    }
                } catch (e) {}
            }
            const sku = pg ? String(pg.sku || '') : '';
            const desc = pg ? (pg.description || pg.name || '') : '';
            const gtin = pg ? String(pg.gtin || pg.gtin13 || '') : '';
            const found = new Set();
            if (sku) {
                document.querySelectorAll('img, source, a').forEach(el => {
                    ['src','data-src','data-zoom-image','data-image','srcset','href'].forEach(attr => {
                        const v = el.getAttribute && el.getAttribute(attr);
                        if (v && v.includes('/imagens/product/' + sku + '/')) {
                            v.split(',').forEach(part => {
                                const u = part.trim().split(' ')[0];
                                if (u.includes('/imagens/product/' + sku + '/')) found.add(u);
                            });
                        }
                    });
                });
            }
            if (pg && pg.image) (Array.isArray(pg.image) ? pg.image : [pg.image]).forEach(u => found.add(u));
            const unique = new Map();
            [...found].forEach(u => {
                const m = u.match(/\\/product\\/\\d+\\/([a-f0-9]{8}-[a-f0-9-]+?)-[a-z]/i);
                const key = m ? m[1] : u;
                if (!unique.has(key)) unique.set(key, u);
            });
            return { sku, gtin, desc, imgs: [...unique.values()] };
        }""")

        if not dados or (not dados.get("desc") and not dados.get("imgs")):
            print(f"  – BLZ: Sem dados para EAN {ean}")
            return "", []

        imgs = [_blz_normalizar_url(u) for u in dados.get("imgs", []) if u]
        desc = dados.get("desc", "")
        print(f"  [BLZ] ✓ SKU {dados.get('sku','')} | {len(imgs)} imagens")
        return desc, imgs

    except Exception as e:
        print(f"  ✗ BLZ: Erro para EAN {ean}: {e}")
        return "", []


# ─── PROCESSAMENTO PRINCIPAL ──────────────────────────────────────────────────

def processar_planilha():
    print("=" * 60)
    print("  Automação de Coleta - Farmácias Online")
    print("=" * 60)

    if not Path(PLANILHA_ENTRADA).exists():
        print(f"✗ Planilha '{PLANILHA_ENTRADA}' não encontrada!")
        return

    wb  = openpyxl.load_workbook(PLANILHA_ENTRADA)
    ws  = wb.active

    col_ean    = col_letter_to_index(COLUNA_EAN)
    col_dsp    = col_letter_to_index(COLUNA_DESC_DSP)
    col_panvel = col_letter_to_index(COLUNA_DESC_PANVEL)
    col_blz    = col_letter_to_index(COLUNA_DESC_BLZ)   # <-- NOVO

    # Cabeçalhos (se vazios)
    if not ws.cell(row=1, column=col_dsp).value:
        ws.cell(row=1, column=col_dsp).value    = "Descrição Drogaria SP"
    if not ws.cell(row=1, column=col_panvel).value:
        ws.cell(row=1, column=col_panvel).value  = "Descrição Panvel"
    if not ws.cell(row=1, column=col_blz).value:                      # <-- NOVO
        ws.cell(row=1, column=col_blz).value     = "Descrição Beleza na Web"

    # Coleta os EANs
    eans = []
    for row in ws.iter_rows(min_row=2):
        cell = row[col_ean - 1]
        if cell.value is not None:
            eans.append((cell.row, str(cell.value).strip()))

    print(f"\n📋 {len(eans)} EAN(s) encontrado(s).\n")

    http_session = requests.Session()
    http_session.headers.update(HEADERS)

    # Inicia o Playwright (um browser para todos os EANs — mais eficiente)
    with sync_playwright() as pw:
        # headless=False ajuda MUITO contra o Akamai. Se precisar rodar sem tela,
        # use headless=True com o pacote 'playwright-stealth' (ver nota abaixo).
        browser = pw.chromium.launch(
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process",
            ],
        )
        context = browser.new_context(
            user_agent=HEADERS["User-Agent"],
            locale="pt-BR",
            viewport={"width": 1366, "height": 768},
            extra_http_headers={"Accept-Language": "pt-BR,pt;q=0.9"},
        )
        # Remove o sinal navigator.webdriver=true que o Akamai detecta
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
        )
        page = context.new_page()

        for idx, (row_num, ean) in enumerate(eans, 1):
            print(f"\n[{idx}/{len(eans)}] EAN: {ean}")
            print("-" * 50)

            # ── Drogaria São Paulo ─────────────────────────────────────────
            desc_dsp, imgs_dsp = coletar_dsp(ean, http_session)
            ws.cell(row=row_num, column=col_dsp).value = limpar_html(desc_dsp)

            # ── Panvel ────────────────────────────────────────────────────
            desc_panvel, imgs_panvel = coletar_panvel(ean, page)
            ws.cell(row=row_num, column=col_panvel).value = limpar_html(desc_panvel)

            # ── Beleza na Web ─────────────────────────────────────────────  <-- NOVO
            desc_blz, imgs_blz = coletar_belezanaweb(ean, page)
            ws.cell(row=row_num, column=col_blz).value = limpar_html(desc_blz)

            # ── Imagens (DSP + Panvel + BLZ, sem duplicatas, max 10) ──────
            todas_urls = list(dict.fromkeys(imgs_dsp + imgs_panvel + imgs_blz))  # dedup por URL
            if todas_urls:
                print(f"\n  📥 Baixando imagens ({len(todas_urls)} URLs únicas)...")
                total = salvar_imagens_unicas(ean, todas_urls, http_session)
                print(f"  ✅ {total} imagem(ns) salva(s) em '{PASTA_IMAGENS}/{ean}/'")
            else:
                print(f"  ⚠ Nenhuma imagem encontrada para EAN {ean}")

            # Salva a planilha após cada EAN (segurança)
            wb.save(PLANILHA_ENTRADA)
            time.sleep(1.5)

        browser.close()

    print("\n" + "=" * 60)
    print(f"✅ Concluído! Planilha salva: {PLANILHA_ENTRADA}")
    print(f"📁 Imagens em: {Path(PASTA_IMAGENS).resolve()}")
    print("=" * 60)


if __name__ == "__main__":
    processar_planilha()